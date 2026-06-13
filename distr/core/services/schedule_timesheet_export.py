"""Excel timesheet export for schedule calendar blocks."""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Optional

UNASSIGNED_BOARD_KEY = "__unassigned__"


def board_key_for_block(block: Any) -> str:
    provider = (getattr(block, "board_provider", None) or "local").strip().lower()
    board_id = getattr(block, "board_id", None)
    external_board_id = (getattr(block, "external_board_id", None) or "").strip()
    if provider == "local" and board_id:
        return f"local:{int(board_id)}"
    if provider != "local":
        external = external_board_id or (str(board_id).strip() if board_id else "")
        if external:
            return f"{provider}:{external}"
    return UNASSIGNED_BOARD_KEY


def duration_minutes(start_at: datetime, end_at: datetime) -> int:
    seconds = max(0, int((end_at - start_at).total_seconds()))
    return max(1, round(seconds / 60))


def format_duration_label(minutes: int) -> str:
    hours, mins = divmod(max(0, int(minutes)), 60)
    if hours and mins:
        return f"{hours}h {mins}m"
    if hours:
        return f"{hours}h"
    return f"{mins}m"


def build_timesheet_workbook(rows: list[dict[str, Any]], *, period_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1A237E")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    project_fill = PatternFill("solid", fgColor="F97316")
    project_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="F3F4F6")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = f"Schedule timesheet — {period_label}"
    title_cell.font = Font(bold=True, size=14, color="111827")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    headers = [
        "Date",
        "Start",
        "End",
        "Duration",
        "Title",
        "Description",
        "Ticket",
        "Board",
    ]
    header_row = 3
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[header_row].height = 20

    current_row = header_row + 1
    current_project: Optional[str] = None
    alt = False

    for entry in rows:
        project_name = entry.get("project_name") or "No project"
        if project_name != current_project:
            current_project = project_name
            alt = False
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
            project_cell = ws.cell(row=current_row, column=1, value=f"Project: {project_name}")
            project_cell.fill = project_fill
            project_cell.font = project_font
            project_cell.alignment = Alignment(horizontal="left", vertical="center")
            project_cell.border = thin_border
            ws.row_dimensions[current_row].height = 20
            current_row += 1

        values = [
            entry.get("date"),
            entry.get("start_time"),
            entry.get("end_time"),
            entry.get("duration_label"),
            entry.get("title"),
            entry.get("description"),
            entry.get("ticket"),
            entry.get("board_name"),
        ]
        row_fill = alt_fill if alt else PatternFill()
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=col, value=value or "")
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=col in {5, 6})
        alt = not alt
        current_row += 1

    widths = [12, 10, 10, 11, 28, 34, 16, 22]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A{header_row}:H{max(header_row, current_row - 1)}"

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def timesheet_filename(start_date: str, end_date: str) -> str:
    safe_start = (start_date or "start")[:10]
    safe_end = (end_date or "end")[:10]
    return f"timesheet_{safe_start}_to_{safe_end}.xlsx"


def parse_export_dates(start_date: str, end_date: str) -> tuple[datetime, datetime, str]:
    start_day = datetime.fromisoformat(str(start_date).strip()[:10]).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    end_day = datetime.fromisoformat(str(end_date).strip()[:10]).replace(
        hour=0, minute=0, second=0, microsecond=0
    )
    if end_day < start_day:
        end_day = start_day
    query_end = end_day + timedelta(days=1)
    period_label = (
        f"{start_day.strftime('%b %d, %Y')} – {end_day.strftime('%b %d, %Y')}"
        if start_day.date() != end_day.date()
        else start_day.strftime("%b %d, %Y")
    )
    return start_day, query_end, period_label
